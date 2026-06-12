from pathlib import Path
from collections import Counter
from openpyxl import load_workbook
import re

SPECIES_DIR = Path("species/helianthus_annuus")

xlsx_file = SPECIES_DIR / "data/raw/pone.0051360.s006.xlsx"
metadata_out = SPECIES_DIR / "data/metadata/sunflower_features_metadata.tsv"
fasta_out = SPECIES_DIR / "data/markers/sunflower_features_25bp.fasta"
summary_out = SPECIES_DIR / "results/qc/sunflower_features_dataset_summary.txt"

metadata_out.parent.mkdir(parents=True, exist_ok=True)
fasta_out.parent.mkdir(parents=True, exist_ok=True)
summary_out.parent.mkdir(parents=True, exist_ok=True)

expected_header = [
    "Chip X",
    "Chip Y",
    "Unigene",
    "Linkage group",
    "centiMorgan position",
    "Plants scored",
    "Mis-matches to template",
    "Sequence",
]

wb = load_workbook(xlsx_file, read_only=True, data_only=True)
ws = wb["Sheet1"]

header = [cell.value for cell in ws[2]][:8]
if header != expected_header:
    raise ValueError(f"Unexpected header:\n{header}")

rows = []
bad_sequences = []

for row in ws.iter_rows(min_row=3, values_only=True):
    chip_x, chip_y, unigene, lg, cm, plants_scored, mismatches_to_template, seq = row[:8]

    if chip_x is None:
        continue

    chip_x = str(chip_x).strip()
    chip_y = str(chip_y).strip()
    unigene = str(unigene).strip()
    lg = str(lg).strip()
    cm = str(cm).strip()
    plants_scored = str(plants_scored).strip()
    mismatches_to_template = str(mismatches_to_template).strip()
    seq = str(seq).strip().upper()

    feature_id = f"HA_X{chip_x}_Y{chip_y}"

    if len(seq) != 25 or not re.fullmatch(r"[ACGT]+", seq):
        bad_sequences.append((feature_id, seq))

    rows.append({
        "feature_id": feature_id,
        "chip_x": chip_x,
        "chip_y": chip_y,
        "unigene": unigene,
        "linkage_group": lg,
        "cM": cm,
        "plants_scored": plants_scored,
        "mismatches_to_template": mismatches_to_template,
        "sequence": seq,
    })

with open(metadata_out, "w", encoding="utf-8") as out:
    out.write("\t".join([
        "feature_id",
        "chip_x",
        "chip_y",
        "unigene",
        "linkage_group",
        "cM",
        "plants_scored",
        "mismatches_to_template",
        "sequence",
    ]) + "\n")

    for r in rows:
        out.write("\t".join([
            r["feature_id"],
            r["chip_x"],
            r["chip_y"],
            r["unigene"],
            r["linkage_group"],
            r["cM"],
            r["plants_scored"],
            r["mismatches_to_template"],
            r["sequence"],
        ]) + "\n")

with open(fasta_out, "w", encoding="utf-8") as out:
    for r in rows:
        out.write(f">{r['feature_id']}\n{r['sequence']}\n")

lg_counts = Counter(r["linkage_group"] for r in rows)
seq_len_counts = Counter(len(r["sequence"]) for r in rows)
template_mismatch_counts = Counter(r["mismatches_to_template"] for r in rows)

with open(summary_out, "w", encoding="utf-8") as out:
    out.write("Sunflower PLOS ONE Dataset S6 summary\n")
    out.write("=====================================\n\n")
    out.write(f"Input file: {xlsx_file}\n")
    out.write(f"Total feature rows: {len(rows)}\n")
    out.write(f"Unique feature IDs: {len(set(r['feature_id'] for r in rows))}\n")
    out.write(f"Unique sequences: {len(set(r['sequence'] for r in rows))}\n")
    out.write(f"Unique unigenes: {len(set(r['unigene'] for r in rows))}\n")
    out.write(f"Bad sequences: {len(bad_sequences)}\n\n")

    out.write("Sequence length distribution:\n")
    for k, v in sorted(seq_len_counts.items()):
        out.write(f"{k}\t{v}\n")

    out.write("\nFeatures by linkage group:\n")
    for lg, count in sorted(lg_counts.items(), key=lambda x: int(x[0])):
        out.write(f"LG{lg}\t{count}\n")

    out.write("\nMis-matches to template distribution:\n")
    for mm, count in sorted(template_mismatch_counts.items(), key=lambda x: int(x[0])):
        out.write(f"{mm}\t{count}\n")

print("Done.")
print(f"Metadata: {metadata_out}")
print(f"FASTA: {fasta_out}")
print(f"Summary: {summary_out}")
print(f"Total features: {len(rows)}")
print(f"Bad sequences: {len(bad_sequences)}")
